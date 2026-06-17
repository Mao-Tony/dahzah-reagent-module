"""质量检验试剂/标准品管理 API

提供质量检验模块试剂/标准品台账的增删改查和AI标签识别功能。
"""

import uuid
from datetime import date, timedelta
from typing import Optional, Any
from io import BytesIO
from urllib.parse import quote

from fastapi import APIRouter, Depends, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from app.platform.database import get_db_session
from app.modules.quality.reagent_schemas import (
    CreateReagentRequest,
    UpdateReagentRequest,
    ReagentResponse,
)
from app.modules.quality.reagent_service import ReagentService
from app.core.storage import save_upload_files

router = APIRouter(prefix="/reagent", tags=["质量检验-试剂管理"])


def dict_to_response(data: dict[str, Any]) -> ReagentResponse:
    """将字典转换为响应模型"""
    return ReagentResponse(
        id=str(data.get("id", "")),
        reagent_label_urls=data.get("reagent_label_urls"),
        reagent_name=data.get("reagent_name", ""),
        arrival_date=data.get("arrival_date", date.today()),
        production_date=data.get("production_date"),
        lot_no=data.get("lot_no", ""),
        incoming_lot_no=data.get("incoming_lot_no"),
        expiration_date=data.get("expiration_date", date.today()),
        specification=data.get("specification"),
        category=data.get("category", "reagent"),
        reagent_no=data.get("reagent_no"),
        content=data.get("content"),
        manufacturer=data.get("manufacturer"),
        quantity=float(data.get("quantity", 0)),
        unit=data.get("unit", "g"),
        status=data.get("status", "available"),
        created_by=data.get("created_by"),
        created_at=data.get("created_at"),
        updated_at=data.get("updated_at"),
    )


# ============ API Endpoints ============

@router.post("/recognize", response_model=dict, summary="AI识别试剂标签图片")
async def recognize_reagent_label(
    files: list[UploadFile] = File(..., description="试剂标签图片（支持多张）"),
    session: AsyncSession = Depends(get_db_session),
):
    """上传试剂标签图片，通过AI识别提取试剂信息"""
    if not files:
        return {
            "code": 400,
            "message": "请上传至少一张图片",
            "data": None,
        }

    service = ReagentService(session)

    # 保存上传的图片
    try:
        image_urls = await save_upload_files(files)
    except Exception as e:
        return {
            "code": 500,
            "message": f"图片上传失败: {str(e)}",
            "data": None,
        }

    # 调用AI识别
    try:
        result = await service.recognize_label(image_urls, operator="system")

        return {
            "code": 200,
            "message": "success",
            "data": {
                "reagent_name": result.get("reagent_name"),
                "lot_no": result.get("lot_no"),
                "content": result.get("content"),
                "manufacturer": result.get("manufacturer"),
                "production_date": result.get("production_date"),
                "expiration_date": result.get("expiration_date"),
                "specification": result.get("specification"),
                "confidence": result.get("confidence", 0.0),
            },
        }
    except Exception as e:
        return {
            "code": 500,
            "message": f"AI识别失败: {str(e)}",
            "data": None,
        }


@router.get("/next-lot-no", response_model=dict, summary="获取下一个入场批号")
async def get_next_incoming_lot_no(
    date_str: Optional[str] = Query(None, description="日期(YYYY-MM-DD)，默认今天"),
    session: AsyncSession = Depends(get_db_session),
):
    """获取当天的下一个入场批号
    格式：年月日9序号，例如 260609901 表示2026年6月9日第1个
    """
    service = ReagentService(session)

    target_date = None
    if date_str:
        from datetime import datetime
        try:
            target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            pass

    try:
        next_lot_no = await service.get_next_incoming_lot_no(target_date)
        return {
            "code": 200,
            "message": "success",
            "data": {
                "incoming_lot_no": next_lot_no,
            },
        }
    except Exception as e:
        return {
            "code": 500,
            "message": f"获取入场批号失败: {str(e)}",
            "data": None,
        }


@router.get("/list", response_model=dict, summary="获取试剂台账列表")
async def get_reagent_list(
    keyword: Optional[str] = Query(None, description="关键词搜索"),
    category: Optional[str] = Query(None, description="分类（试剂/标准品）"),
    status: Optional[str] = Query(None, description="状态"),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=100, description="每页数量"),
    session: AsyncSession = Depends(get_db_session),
):
    """获取质量检验试剂/标准品台账列表"""
    service = ReagentService(session)

    reagents, total = await service.list_reagents(
        keyword=keyword,
        category=category,
        status=status,
        page=page,
        page_size=page_size,
    )

    items = [dict_to_response(r) for r in reagents]

    return {
        "code": 200,
        "message": "success",
        "data": {
            "items": [item.model_dump() for item in items],
            "total": total,
            "page": page,
            "page_size": page_size,
        },
    }


@router.get("/export", summary="导出试剂台账Excel")
async def export_reagents_excel(
    keyword: Optional[str] = Query(None, description="关键词搜索"),
    category: Optional[str] = Query(None, description="分类"),
    status: Optional[str] = Query(None, description="状态"),
    session: AsyncSession = Depends(get_db_session),
):
    """导出试剂/标准品台账为Excel文件"""
    service = ReagentService(session)

    # 获取所有数据（不分页）
    reagents, _ = await service.list_reagents(
        keyword=keyword,
        category=category,
        status=status,
        page=1,
        page_size=10000,
    )

    # 创建Excel工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "试剂台账"

    # 定义样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 表头
    headers = [
        "试剂名称", "到货日期", "生产日期", "批号", "入场批号",
        "有效期", "规格", "分类", "编号", "含量", "生产厂家",
        "数量", "单位", "状态", "创建人", "创建时间"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 分类选项
    category_labels = {
        "reagent": "试剂",
        "standard": "标准品",
        "reference": "对照品",
        "consumable": "耗材",
    }
    status_labels = {
        "available": "可用",
        "low_stock": "库存不足",
        "expired": "已过期",
        "quarantine": "待检",
        "scrap": "报废",
    }

    # 填充数据
    for row_idx, reagent in enumerate(reagents, 2):
        reagent_dict = dict_to_response(reagent).model_dump()

        row_data = [
            reagent_dict.get("reagent_name", ""),
            str(reagent_dict.get("arrival_date", "")) if reagent_dict.get("arrival_date") else "",
            str(reagent_dict.get("production_date", "")) if reagent_dict.get("production_date") else "",
            reagent_dict.get("lot_no", ""),
            reagent_dict.get("incoming_lot_no", ""),
            str(reagent_dict.get("expiration_date", "")) if reagent_dict.get("expiration_date") else "",
            reagent_dict.get("specification", ""),
            category_labels.get(reagent_dict.get("category", ""), reagent_dict.get("category", "")),
            reagent_dict.get("reagent_no", ""),
            reagent_dict.get("content", ""),
            reagent_dict.get("manufacturer", ""),
            str(reagent_dict.get("quantity", "")),
            reagent_dict.get("unit", ""),
            status_labels.get(reagent_dict.get("status", ""), reagent_dict.get("status", "")),
            reagent_dict.get("created_by", ""),
            str(reagent_dict.get("created_at", "")) if reagent_dict.get("created_at") else "",
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # 自动调整列宽
    for col in range(1, len(headers) + 1):
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col)
        column = ws.column_dimensions[column_letter]
        for row in ws.iter_rows():
            cell = row[col - 1]
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        column.width = adjusted_width

    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"试剂台账_{date.today().strftime('%Y%m%d')}.xlsx"
    encoded_filename = quote(filename)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


@router.get("/{reagent_id}", response_model=dict, summary="获取试剂详情")
async def get_reagent_detail(
    reagent_id: str,
    session: AsyncSession = Depends(get_db_session),
):
    """根据ID获取试剂详情"""
    service = ReagentService(session)
    reagent = await service.get_reagent_by_id(reagent_id)

    if not reagent:
        return {
            "code": 404,
            "message": "试剂记录不存在",
            "data": None,
        }

    return {
        "code": 200,
        "message": "success",
        "data": dict_to_response(reagent).model_dump(),
    }


@router.post("", response_model=dict, summary="创建试剂记录")
async def create_reagent(
    request: CreateReagentRequest,
    session: AsyncSession = Depends(get_db_session),
):
    """创建新的试剂记录"""
    service = ReagentService(session)

    # 处理生产日期：如果未提供，默认到货日期+3年
    production_date = request.production_date
    if not production_date:
        production_date = request.arrival_date + timedelta(days=3 * 365)

    try:
        reagent = await service.create_reagent(
            data={
                "reagent_label_urls": request.reagent_label_urls,
                "reagent_name": request.reagent_name,
                "arrival_date": request.arrival_date,
                "production_date": production_date,
                "lot_no": request.lot_no,
                "incoming_lot_no": request.incoming_lot_no,
                "expiration_date": request.expiration_date,
                "specification": request.specification,
                "category": request.category,
                "reagent_no": request.reagent_no,
                "content": request.content,
                "manufacturer": request.manufacturer,
                "quantity": request.quantity,
                "unit": request.unit,
            },
            operator="system",
        )

        return {
            "code": 200,
            "message": "success",
            "data": dict_to_response(reagent).model_dump(),
        }
    except Exception as e:
        return {
            "code": 500,
            "message": f"创建失败: {str(e)}",
            "data": None,
        }


@router.put("/{reagent_id}", response_model=dict, summary="更新试剂记录")
async def update_reagent(
    reagent_id: str,
    request: UpdateReagentRequest,
    session: AsyncSession = Depends(get_db_session),
):
    """更新试剂记录"""
    service = ReagentService(session)

    try:
        reagent = await service.update_reagent(
            reagent_id=reagent_id,
            data=request.model_dump(exclude_unset=True),
        )

        if not reagent:
            return {
                "code": 404,
                "message": "试剂记录不存在",
                "data": None,
            }

        return {
            "code": 200,
            "message": "success",
            "data": dict_to_response(reagent).model_dump(),
        }
    except Exception as e:
        return {
            "code": 500,
            "message": f"更新失败: {str(e)}",
            "data": None,
        }


@router.delete("/{reagent_id}", response_model=dict, summary="删除试剂记录")
async def delete_reagent(
    reagent_id: str,
    session: AsyncSession = Depends(get_db_session),
):
    """删除试剂记录"""
    service = ReagentService(session)

    success = await service.delete_reagent(reagent_id)

    if not success:
        return {
            "code": 404,
            "message": "试剂记录不存在",
            "data": None,
        }

    return {
        "code": 200,
        "message": "success",
        "data": None,
    }